import openpyxl
import requests
import re
from htmlparsing import Element, HTMLParsing, Text, Attr, Parse, HTML, Markdown
wb = openpyxl.Workbook()
wb = openpyxl.load_workbook(filename = 'C:\\Users/User/Desktop/book.xlsx')
sheets = wb.sheetnames
sheet = wb[sheets[0]]
c = 1
print('Working...')
for a in range (1,100):
    url = 'https://kontakt.by/pinsk/predprijatija/stroitelstvo-stroitelnie-materiali-steklo-keramika-248/stroitelstvo-250/stroitelnie-i-otdelochnie-raboti-524/?q_l=%D0%9F%D0%B8%D0%BD%D1%81%D0%BA&l%5B0%5D=42&c%5B0%5D=248&c%5B1%5D=250&c%5B2%5D=524&page=' + str(a)
    r = requests.get(url)
    e = Element(text=r.text)
    for i in range(1,26):
        try:
            sheet['A%s' % c]=(e.xpath('//*[@id="companies"]/li[%s]/div/p/a' % i)[0].text).strip()
        except IndexError:
            break
        try:
            sheet['B%s' % c]=(e.xpath('//*[@id="companies"]/li[%s]/div/ul[1]/li[1]/span' % i)[0].text).strip()
        except IndexError:
            sheet['B%s' % c]='Адрес отсутствует'
        
        try:
            sheet['C%s' % c]=re.sub('[\s+]', '', (e.xpath('//*[@id="companies"]/li[%s]/div/ul[1]/li[2]' % i)[0].text).strip())
        except IndexError:
            sheet['C%s' % c]='Номер отсутствует'
        try:
            sheet['D%s' % c]=(e.xpath('//*[@id="companies"]/li[%s]/div/ul[2]/li/a' % i)[0].text).strip()
        except IndexError:
            sheet['D%s' % c]='Категория отсутствует'
        c+=1
wb.save('C:\\Users/User/Desktop/book.xlsx')
print('End.')
